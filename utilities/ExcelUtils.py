import openpyxl
from utilities import ReadConfiguration

def get_data_from_excel():
   global workbook
   try:
      path=ReadConfiguration.read_configuration("Login excel info", "excell_path")
      print("Excel Path is :-",path)
      sheet_name = ReadConfiguration.read_configuration("Login excel info", "sheet_name_Login_with_valid_credentials")
      print("Sheet Name is :-",sheet_name)
      final_list=[]
      workbook= openpyxl.load_workbook(path)
      sheet_names=workbook[sheet_name]
      total_rows=sheet_names.max_row
      total_cols = sheet_names.max_column
      for r in range(2,total_rows+1):
         row_list=[]
         for c in range(1,total_cols+1):
            row_list.append(sheet_names.cell(row=r,column=c).value)
         final_list.append(row_list)

      return final_list
   except Exception as e:
      e.with_traceback()
   finally:
      workbook.close()

